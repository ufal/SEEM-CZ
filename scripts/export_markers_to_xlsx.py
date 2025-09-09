#!/usr/bin/env python3
"""
Script to export marker annotation files to XLSX format for Google Sheets import.
The output follows the schema definition order and includes expression and language columns.
"""

import argparse
import logging
import os
import re
import sys
import xml.etree.ElementTree as ET
from markerdoc import MarkerDoc, MarkerDocDef

# Column names in Czech
EXPRESSION_COLUMN = 'Výraz'
LANGUAGE_COLUMN = 'Jazyk'

# Verbtag position names in Czech (from https://wiki.korpus.cz/doku.php/cnk:syn2020:verbtag)
VERBTAG_POSITIONS = [
    'Typ slovesa',          # Position 1: V, A, or -
    'Typ slovesného tvaru', # Position 2: D, C, I, F, T, O, or -
    'Slovesný rod',         # Position 3: A, P, p, or -
    'Osoba',               # Position 4: 1, 2, 3, or -
    'Číslo',               # Position 5: S, P, v, or -
    'Čas'                  # Position 6: P, F, B, R, Q, or -
]

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    print("Error: openpyxl library is required. Install it with: pip install openpyxl")
    sys.exit(1)

def setup_logging():
    """Setup logging configuration"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )

def extract_expression_and_language(filename):
    """
    Extract expression and language from marker filename.
    
    Expected patterns:
    - markers_gold-cs-asi.xml -> ("asi", "cs")
    - markers_gold-en-urcite.xml -> ("urcite", "en")
    
    Args:
        filename: Name of the marker file
    
    Returns:
        Tuple (expression, language) or (None, None) if pattern not found
    """
    # Remove .xml extension
    base_name = filename.replace('.xml', '')
    
    # Pattern: markers_gold-{lang}-{expression}
    match = re.match(r'markers(?:_gold)?-([a-z]{2})-(.+)', base_name)
    if match:
        language, expression = match.groups()
        return expression, language
    
    # If no pattern matches, try to extract from the end
    # Look for -cs- or -en- pattern
    lang_match = re.search(r'-([a-z]{2})-([^-]+)$', base_name)
    if lang_match:
        language, expression = lang_match.groups()
        return expression, language
    
    logging.warning(f"Could not extract expression and language from filename: {filename}")
    return None, None

def is_verbtag_column(column_name):
    """
    Check if a column contains verbtag data.
    
    Args:
        column_name: Display name of the column
    
    Returns:
        True if the column contains verbtag data, False otherwise
    """
    # Check if the column name contains "verbtag" (case insensitive)
    return 'verbtag' in column_name.lower()

def decompose_verbtag(verbtag_value):
    """
    Decompose a verbtag value into its 6 positions.
    If there are multiple verbtags, select the most informative one (lowest number of hyphens).
    In case of ties, take the first such one.
    
    Args:
        verbtag_value: String containing verbtag(s), potentially space-separated
    
    Returns:
        List of dictionaries, each containing the 6 position values for one verbtag
    """
    if not verbtag_value or not verbtag_value.strip():
        return []
    
    # Split by whitespace to handle multiple verbtags
    verbtags = verbtag_value.strip().split()
    informative_tags = []
    
    for verbtag in verbtags:
        if len(verbtag) == 6:
            # Count hyphens to determine informativeness
            hyphen_count = verbtag.count('-')
            # Skip if the verbtag carries no information (all hyphens)
            if hyphen_count == len(verbtag):
                continue
            
            # Decompose the 6-character verbtag
            positions = {
                VERBTAG_POSITIONS[0]: verbtag[0],  # Typ slovesa
                VERBTAG_POSITIONS[1]: verbtag[1],  # Typ slovesného tvaru
                VERBTAG_POSITIONS[2]: verbtag[2],  # Slovesný rod
                VERBTAG_POSITIONS[3]: verbtag[3],  # Osoba
                VERBTAG_POSITIONS[4]: verbtag[4],  # Číslo
                VERBTAG_POSITIONS[5]: verbtag[5]   # Čas
            }
            informative_tags.append((hyphen_count, verbtag, positions))
    
    if len(informative_tags) > 1:
        # Sort by hyphen count (ascending) - fewer hyphens = more informative
        informative_tags.sort(key=lambda x: x[0])
        most_informative = informative_tags[0]
        
        # Check if there are ties with the same hyphen count
        tied_tags = [tag for tag in informative_tags if tag[0] == most_informative[0]]
        if len(tied_tags) > 1:
            logging.info(f"Multiple verbtags with same informativeness found: {verbtag_value}. "
                        f"Using the first one with {most_informative[0]} hyphens: {most_informative[1]}")
        else:
            logging.info(f"Selected most informative verbtag from: {verbtag_value}. "
                        f"Using: {most_informative[1]} (with {most_informative[0]} hyphens)")
        
        return [most_informative[2]]  # Return the positions dictionary
    elif len(informative_tags) == 1:
        return [informative_tags[0][2]]  # Return the positions dictionary
    else:
        # No informative verbtags found, return empty positions
        return []

def create_verbtag_column_names(base_column_name, tag_index=None):
    """
    Create column names for decomposed verbtag positions.
    
    Args:
        base_column_name: Original column name containing verbtag
        tag_index: Index of the verbtag (for multiple verbtags), None for single verbtag
    
    Returns:
        List of column names for the 6 positions
    """
    if tag_index is not None:
        # For multiple verbtags, add index: "Predikát (verbtag) - Typ slovesa (1)"
        return [f"{base_column_name} - {pos_name} ({tag_index + 1})" for pos_name in VERBTAG_POSITIONS]
    else:
        # For single verbtag: "Predikát (verbtag) - Typ slovesa"
        return [f"{base_column_name} - {pos_name}" for pos_name in VERBTAG_POSITIONS]

def get_column_order(def_doc):
    """
    Get the column order from the definition file using display names,
    including decomposed verbtag columns.
    
    Args:
        def_doc: MarkerDocDef instance
    
    Returns:
        List of column display names in definition order
    """
    columns = [EXPRESSION_COLUMN, LANGUAGE_COLUMN]  # First two columns with display names in Czech
    
    # Add all attributes from the definition in order using display names
    for interp_elem in def_doc.xml.findall(".//interp"):
        key = interp_elem.attrib.get("key")
        if key:
            display_name = def_doc.get_display_string(key)
            columns.append(display_name)
            
            # If this is a verbtag column, add decomposed columns
            if is_verbtag_column(display_name):
                # Add columns for each verbtag position (no need for multiple verbtag handling)
                verbtag_columns = create_verbtag_column_names(display_name)
                columns.extend(verbtag_columns)
    
    return columns

def process_marker_file(file_path, def_doc, expression, language):
    """
    Process a single marker file and return rows for XLSX export.
    
    Args:
        file_path: Path to the marker XML file
        def_doc: MarkerDocDef instance for schema information
        expression: Expression extracted from filename
        language: Language extracted from filename
    
    Returns:
        List of dictionaries representing rows with display names as keys
    """
    logging.info(f"Processing marker file: {file_path}")
    
    try:
        marker_doc = MarkerDoc(file_path)
    except Exception as e:
        logging.error(f"Failed to parse {file_path}: {e}")
        return []
    
    rows = []
    
    for item_elem in marker_doc:
        row = {
            EXPRESSION_COLUMN: expression or '',
            LANGUAGE_COLUMN: language or ''
        }
        
        # Add all attributes from the item element using display names
        for attr_name, attr_value in item_elem.attrib.items():
            display_name = def_doc.get_display_string(attr_name)
            # For values, also try to get display string
            display_value = def_doc.get_display_string(attr_name, attr_value)
            row[display_name] = display_value
            
            # If this is a verbtag column, decompose it
            if is_verbtag_column(display_name):
                decomposed_tags = decompose_verbtag(attr_value)
                
                if len(decomposed_tags) == 1:
                    # Single informative verbtag - add columns without index
                    verbtag_columns = create_verbtag_column_names(display_name)
                    for i, pos_name in enumerate(VERBTAG_POSITIONS):
                        row[verbtag_columns[i]] = decomposed_tags[0][pos_name]
                else:
                    # No informative verbtags - add empty columns
                    verbtag_columns = create_verbtag_column_names(display_name)
                    for col_name in verbtag_columns:
                        row[col_name] = ''
        
        rows.append(row)
    
    logging.info(f"Extracted {len(rows)} items from {file_path}")
    return rows

def export_to_xlsx(marker_files, output_file, schema_file):
    """
    Export marker files to XLSX format.
    
    Args:
        marker_files: List of marker file paths
        output_file: Path to output XLSX file
        schema_file: Path to schema definition XML file
    """
    logging.info(f"Exporting {len(marker_files)} marker files to XLSX")
    logging.info(f"Using schema definition: {schema_file}")
    logging.info(f"Output will be saved to: {output_file}")
    
    # Load the schema definition
    def_doc = MarkerDocDef(schema_file)
    
    all_rows = []
    
    # Process each marker file to collect all data
    for file_path in marker_files:
        filename = os.path.basename(file_path)
        expression, language = extract_expression_and_language(filename)
        
        if expression is None or language is None:
            logging.warning(f"Skipping {file_path} - could not determine expression/language")
            continue
        
        rows = process_marker_file(file_path, def_doc, expression, language)
        all_rows.extend(rows)
    
    # Get column order from definition
    columns = get_column_order(def_doc)
    logging.info(f"Column order: {columns}")
    
    # Create XLSX file
    logging.info(f"Writing {len(all_rows)} total rows to XLSX file")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Marker Annotations"
    
    # Write header row with bold formatting
    header_font = Font(bold=True)
    for col_idx, column_display_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column_display_name)
        cell.font = header_font
    
    # Write data rows
    for row_idx, row_data in enumerate(all_rows, 2):
        for col_idx, column_display_name in enumerate(columns, 1):
            value = row_data.get(column_display_name, '')
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Set a reasonable maximum width
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(output_file)
    logging.info(f"XLSX export completed successfully: {output_file}")

def validate_marker_files(file_paths):
    """
    Validate that marker files exist.
    
    Args:
        file_paths: List of marker file paths
    
    Returns:
        List of valid marker file paths
    """
    valid_files = []
    
    for file_path in file_paths:
        if os.path.exists(file_path):
            valid_files.append(file_path)
        else:
            logging.error(f"File not found: {file_path}")
    
    logging.info(f"Validated {len(valid_files)} out of {len(file_paths)} marker files")
    return valid_files

def main():
    """Main function to handle command line arguments and process files"""
    parser = argparse.ArgumentParser(
        description="Export marker annotation files to XLSX format for Google Sheets import"
    )
    parser.add_argument(
        "input_files",
        nargs='+',
        help="List of marker XML files to process"
    )
    parser.add_argument(
        "-o", "--output",
        required=True,
        help="Output XLSX file path"
    )
    parser.add_argument(
        "-s", "--schema",
        default="teitok/config/markers_def.xml",
        help="Schema definition file (default: teitok/config/markers_def.xml)"
    )
    
    args = parser.parse_args()
    
    setup_logging()
    
    # Validate schema file exists
    if not os.path.exists(args.schema):
        logging.error(f"Schema file not found: {args.schema}")
        sys.exit(1)
    
    # Validate marker files
    marker_files = validate_marker_files(args.input_files)
    
    if not marker_files:
        logging.error("No valid marker files found")
        sys.exit(1)
    
    # Ensure output directory exists
    output_dir = os.path.dirname(args.output)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    try:
        export_to_xlsx(marker_files, args.output, args.schema)
        logging.info("Export completed successfully")
    except Exception as e:
        logging.error(f"Error during export: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
